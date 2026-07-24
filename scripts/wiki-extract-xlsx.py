#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
wiki-extract-xlsx.py — extract markdown from an Excel workbook via openpyxl.

Stage 12 of the wiki upgrade arc — see [[Eric Ma — Mastering PKM with Obsidian
and AI]] for the canonical recipe. Per Ma: openpyxl, NOT pandas. Real
spreadsheets are messy (merged cells, free text, arbitrary layouts); pandas
assumes tabular shape. Progressive reveal is the right pattern: first map the
spreadsheet's architecture, then drill into specific sheets/ranges.

Usage:
  uv run wiki-extract-xlsx.py <path>                                # schema map (default)
  uv run wiki-extract-xlsx.py <path> --list-sheets                   # sheet names only
  uv run wiki-extract-xlsx.py <path> --sheet "<name>"                # sheet's used range as table
  uv run wiki-extract-xlsx.py <path> --sheet "<name>" --range A1:E20 # exact range
  uv run wiki-extract-xlsx.py --self-test                             # round-trip exit 0/1

LIMITATIONS:
- Formulas evaluated lazily; `data_only=True` uses last-cached values.
  Open the workbook in Excel once and re-save if cached values are stale.
- Charts and embedded images not extracted.
- Workbooks with >10 sheets: schema map truncates to first 10. Use
  `--list-sheets` for the full list.

First run cost: ~10-30 sec (openpyxl is pure Python, fast resolve). Cached after.
"""
import sys
import tempfile
from pathlib import Path

# Force UTF-8 on stdout/stderr so non-ASCII content (accented chars, em-dashes,
# math symbols) doesn't crash on Windows where the default codec is cp1252.
# Python 3.7+ feature; harmless on Linux/macOS.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "missing dep: openpyxl. Run via `uv run wiki-extract-xlsx.py ...` "
        "(PEP-723 inline deps), not bare python.\n"
    )
    sys.exit(1)


SCHEMA_TRUNC = 10  # truncate schema map at N sheets


def parse_args(argv):
    args = {"path": None, "sheet": None, "range": None, "list_sheets": False, "self_test": False}
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--self-test":
            args["self_test"] = True
        elif a == "--list-sheets":
            args["list_sheets"] = True
        elif a == "--sheet" and i + 1 < len(argv):
            args["sheet"] = argv[i + 1]; i += 1
        elif a.startswith("--sheet="):
            args["sheet"] = a[len("--sheet="):]
        elif a == "--range" and i + 1 < len(argv):
            args["range"] = argv[i + 1]; i += 1
        elif a.startswith("--range="):
            args["range"] = a[len("--range="):]
        elif args["path"] is None and not a.startswith("-"):
            args["path"] = a
        i += 1
    return args


def _cell_str(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("|", "\\|")


def _count_nonempty(ws):
    n = 0
    for row in ws.iter_rows(values_only=True):
        n += sum(1 for v in row if v is not None and str(v).strip())
    return n


def _sample_rows(ws, n=5):
    sample = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            sample.append(row)
            if len(sample) >= n:
                break
    return sample


def _defined_names_count(wb):
    try:
        return sum(1 for _ in wb.defined_names)
    except Exception:
        try:
            return len(list(wb.defined_names.definedName))  # legacy
        except Exception:
            return 0


def schema_map(wb, path):
    out = [f"# Spreadsheet: {path.name}", ""]
    out.append("## Workbook structure")
    out.append(f"- Sheets: {len(wb.sheetnames)}")
    sheets_to_show = wb.sheetnames[:SCHEMA_TRUNC]
    total = sum(_count_nonempty(wb[s]) for s in sheets_to_show)
    suffix = f" (first {len(sheets_to_show)} sheets)" if len(wb.sheetnames) > SCHEMA_TRUNC else ""
    out.append(f"- Total cells with content{suffix}: ~{total}")
    out.append(f"- Defined names: {_defined_names_count(wb)}")
    out.append("")

    for name in sheets_to_show:
        ws = wb[name]
        out.append(f"## Sheet: {name}")
        out.append(f"- Dimensions: {ws.dimensions} (used range)")
        out.append(f"- Merged cells: {len(ws.merged_cells.ranges)}")
        sample = _sample_rows(ws, 5)
        if sample:
            cols = max(len(r) for r in sample)
            header = "| " + " | ".join(get_column_letter(i + 1) for i in range(cols)) + " |"
            sep = "| " + " | ".join(["---"] * cols) + " |"
            out.append(f"- First {len(sample)} non-empty rows:")
            out.append("  " + header)
            out.append("  " + sep)
            for r in sample:
                cells = [_cell_str(r[i] if i < len(r) else None) for i in range(cols)]
                out.append("  | " + " | ".join(cells) + " |")
        else:
            out.append("- _(empty sheet)_")
        out.append("")

    if len(wb.sheetnames) > SCHEMA_TRUNC:
        more = len(wb.sheetnames) - SCHEMA_TRUNC
        out.append(
            f"_({more} more sheet(s) — pass `--list-sheets` for the full list "
            "or `--sheet <name>` to drill into one.)_"
        )

    return "\n".join(out)


def list_sheets(wb):
    return "\n".join(wb.sheetnames)


def sheet_table(ws, cell_range=None):
    if cell_range:
        cells = ws[cell_range]
        # ws[range] returns a tuple-of-tuples-of-Cells, OR a single Cell if range == "A1"
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        rows = [[_cell_str(c.value) for c in row] for row in cells]
    else:
        rows = [[_cell_str(v) for v in row] for row in ws.iter_rows(values_only=True)]

    if not rows:
        return f"## Sheet: {ws.title}\n\n_(empty)_"

    cols = max(len(r) for r in rows)
    rows = [r + [""] * (cols - len(r)) for r in rows]
    out = [f"## Sheet: {ws.title}"]
    if cell_range:
        out.append(f"_Range: {cell_range}_")
    else:
        out.append(f"_Used range: {ws.dimensions}_")
    out.append("")
    out.append("| " + " | ".join(get_column_letter(i + 1) for i in range(cols)) + " |")
    out.append("| " + " | ".join(["---"] * cols) + " |")
    for r in rows:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def self_test():
    KNOWN = "Stage 12 xlsx selftest"
    KNOWN_VAL = 42
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = Path(f.name)
    try:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "header"
        ws1["B1"] = "value"
        ws1["A2"] = KNOWN
        ws1["B2"] = KNOWN_VAL
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "second sheet marker"
        wb.save(str(tmp_path))

        rt = openpyxl.load_workbook(str(tmp_path), data_only=True)
        smap = schema_map(rt, tmp_path)
        if KNOWN not in smap:
            sys.stderr.write(f"selftest FAIL (schema map): marker not in output\n{smap}\n")
            sys.exit(1)
        if "## Sheet: Sheet1" not in smap or "## Sheet: Sheet2" not in smap:
            sys.stderr.write(f"selftest FAIL: both sheet headers expected\n{smap}\n")
            sys.exit(1)

        tbl = sheet_table(rt["Sheet1"])
        if KNOWN not in tbl or str(KNOWN_VAL) not in tbl:
            sys.stderr.write(f"selftest FAIL (sheet drill): expected marker + {KNOWN_VAL}\n{tbl}\n")
            sys.exit(1)

        rng = sheet_table(rt["Sheet1"], "A1:B2")
        if KNOWN not in rng or "header" not in rng:
            sys.stderr.write(f"selftest FAIL (range drill): expected marker + header\n{rng}\n")
            sys.exit(1)

        sheets = list_sheets(rt)
        if "Sheet1" not in sheets or "Sheet2" not in sheets:
            sys.stderr.write(f"selftest FAIL (list-sheets): missing sheets\n{sheets}\n")
            sys.exit(1)

        sys.stdout.write("selftest pass\n")
        sys.exit(0)
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


def main():
    args = parse_args(sys.argv[1:])
    if args["self_test"]:
        return self_test()
    if not args["path"]:
        sys.stderr.write(
            "usage: uv run wiki-extract-xlsx.py <path> "
            "[--list-sheets | --sheet <name> [--range A1:E20]] | --self-test\n"
        )
        sys.exit(2)

    path = Path(args["path"])
    if not path.exists():
        sys.stderr.write(f"file not found: {path}\n")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        sys.stderr.write(f"openpyxl failed: {type(e).__name__}: {e}\n")
        sys.exit(1)

    if args["list_sheets"]:
        sys.stdout.write(list_sheets(wb) + "\n")
    elif args["sheet"]:
        if args["sheet"] not in wb.sheetnames:
            sys.stderr.write(
                f"sheet not found: {args['sheet']}\n  available: {', '.join(wb.sheetnames)}\n"
            )
            sys.exit(1)
        ws = wb[args["sheet"]]
        sys.stdout.write(sheet_table(ws, args["range"]) + "\n")
    else:
        sys.stdout.write(schema_map(wb, path) + "\n")


if __name__ == "__main__":
    main()
